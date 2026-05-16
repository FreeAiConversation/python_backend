from flask import Blueprint, request, jsonify
import tempfile
import os
import base64
from werkzeug.utils import secure_filename
import traceback
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

pdf_to_excel_bp = Blueprint('pdf_to_excel', __name__)

@pdf_to_excel_bp.route('/api/pdf-to-excel', methods=['POST'])
def convert_pdf_to_excel():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded."}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename.lower().endswith('.pdf'):
        return jsonify({"success": False, "error": "Invalid file format. Please upload a PDF file."}), 400

    filename = secure_filename(file.filename)
    base_filename = os.path.splitext(filename)[0]
    
    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = os.path.join(temp_dir, filename)
        output_filename = f"{base_filename}_converted.xlsx"
        output_path = os.path.join(temp_dir, output_filename)

        try:
            file.save(input_path)
            
            with pdfplumber.open(input_path) as pdf:
                all_data = []
                num_cols_detected = 0
                
                for page in pdf.pages:
                    words = page.extract_words()
                    if not words: continue

                    # 1. Dynamic Column Detection
                    # We look at the x-coordinates of words to find clusters (gutters)
                    x_coords = sorted([w["x0"] for w in words])
                    gutters = []
                    if x_coords:
                        curr_min = x_coords[0]
                        for i in range(1, len(x_coords)):
                            # 15px gap suggests a new column in most standard documents
                            if x_coords[i] - x_coords[i-1] > 15:
                                gutters.append((curr_min, x_coords[i-1]))
                                curr_min = x_coords[i]
                        gutters.append((curr_min, x_coords[-1]))

                    num_cols_detected = max(num_cols_detected, len(gutters))

                    # 2. Row Snapping (Premium Grid Logic)
                    rows_raw = {}
                    for w in words:
                        y = round(w["top"] / 8) * 8 # snap to 8px grid
                        rows_raw.setdefault(y, []).append(w)

                    sorted_ys = sorted(rows_raw.keys())
                    page_rows = []
                    pending_row = None
                    
                    for y in sorted_ys:
                        wds = sorted(rows_raw[y], key=lambda w: w["x0"])
                        row_cells = [""] * len(gutters)
                        has_leading_content = False

                        for w in wds:
                            # Map word to detected gutter
                            for idx, (xmin, xmax) in enumerate(gutters):
                                if xmin - 5 <= w["x0"] <= xmax + 15:
                                    sep = " " if row_cells[idx] else ""
                                    row_cells[idx] += sep + w["text"]
                                    if idx == 0: has_leading_content = True
                                    break

                        if has_leading_content:
                            if pending_row: page_rows.append(pending_row)
                            pending_row = row_cells
                        elif pending_row:
                            # Continuation of previous row (multi-line)
                            for i in range(len(gutters)):
                                if row_cells[i]:
                                    sep = "\n" if pending_row[i] else ""
                                    pending_row[i] += sep + row_cells[i]
                    
                    if pending_row: page_rows.append(pending_row)
                    all_data.extend(page_rows)

            # Write Excel with Dynamic Styling
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            # Styles
            header_font = Font(name="Arial", bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", start_color="2F5496")
            alt_fill = PatternFill("solid", start_color="DCE6F1")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            border = Border(left=Side(style="thin", color="AAAAAA"), right=Side(style="thin", color="AAAAAA"), top=Side(style="thin", color="AAAAAA"), bottom=Side(style="thin", color="AAAAAA"))

            for r_idx, row in enumerate(all_data, start=1):
                is_header = r_idx == 1
                is_alt = r_idx % 2 == 0
                max_lines = 1
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(r_idx, c_idx)
                    cell.value = str(val).strip()
                    cell.border = border
                    if is_header:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                    else:
                        cell.alignment = center_align if c_idx > 1 else left_align
                        if is_alt: cell.fill = alt_fill
                    max_lines = max(max_lines, str(val).count("\n") + 1)
                ws.row_dimensions[r_idx].height = max(15, max_lines * 14 + 5)

            # Auto-adjust column widths
            for i in range(1, num_cols_detected + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22

            wb.save(output_path)

            with open(output_path, "rb") as f:
                output_b64 = base64.b64encode(f.read()).decode('utf-8')

            return jsonify({
                "success": True,
                "output_filename": output_filename,
                "output_b64": output_b64,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            })

        except Exception as e:
            traceback.print_exc()
            return jsonify({"success": False, "error": f"Conversion failed: {str(e)}"}), 500
