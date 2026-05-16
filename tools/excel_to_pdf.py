from flask import Blueprint, request, jsonify
import tempfile
import os
import base64
from werkzeug.utils import secure_filename
import traceback
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

excel_to_pdf_bp = Blueprint('excel_to_pdf', __name__)

@excel_to_pdf_bp.route('/api/excel-to-pdf', methods=['POST'])
def convert_excel_to_pdf():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded."}), 400

    file = request.files['file']
    if file.filename == '' or not (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.xls')):
        return jsonify({"success": False, "error": "Invalid file format. Please upload an Excel file."}), 400

    filename = secure_filename(file.filename)
    base_filename = os.path.splitext(filename)[0]
    
    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = os.path.join(temp_dir, filename)
        output_filename = f"{base_filename}_converted.pdf"
        output_path = os.path.join(temp_dir, output_filename)

        try:
            file.save(input_path)
            
            # Load workbook
            wb = openpyxl.load_workbook(input_path, data_only=True)
            ws = wb.active

            # 1. Dynamic Data Extraction
            all_data = []
            max_col = ws.max_column
            for r in range(1, ws.max_row + 1):
                row_data = []
                for c in range(1, max_col + 1):
                    val = ws.cell(r, c).value
                    row_data.append(str(val) if val is not None else "")
                if any(v.strip() for v in row_data):
                    all_data.append(row_data)

            if not all_data:
                return jsonify({"success": False, "error": "The Excel file is empty."}), 400

            # 2. PDF Document Setup
            # If columns > 6, use landscape automatically
            page_size = landscape(letter) if max_col > 6 else letter
            doc = SimpleDocTemplate(
                output_path,
                pagesize=page_size,
                leftMargin=0.4*inch, rightMargin=0.4*inch,
                topMargin=0.4*inch, bottomMargin=0.4*inch,
            )

            styles = getSampleStyleSheet()
            def para(text, bold=False, center=False):
                style = ParagraphStyle(
                    "cell",
                    fontName="Helvetica-Bold" if bold else "Helvetica",
                    fontSize=8 if max_col > 8 else 9,
                    alignment=1 if center else 0,
                    leading=10,
                    textColor=colors.white if bold else colors.black,
                )
                return Paragraph(str(text).replace("\n", "<br/>"), style)

            table_data = []
            for i, row in enumerate(all_data):
                is_header = i == 0
                table_data.append([para(v, bold=is_header, center=True) for v in row])

            # 3. Dynamic Column Widths (proportional)
            total_width = page_size[0] - 0.8*inch
            col_width = total_width / max_col
            col_widths = [col_width] * max_col

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#DCE6F1")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))

            doc.build([tbl])

            with open(output_path, "rb") as f:
                output_b64 = base64.b64encode(f.read()).decode('utf-8')

            return jsonify({
                "success": True,
                "output_filename": output_filename,
                "output_b64": output_b64,
                "mime_type": "application/pdf"
            })

        except Exception as e:
            traceback.print_exc()
            return jsonify({"success": False, "error": f"Conversion failed: {str(e)}"}), 500
